import pandas as pd
import openpyxl
import json, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"D:\西北投资\2026-win\其他工作\制度\附件1-2：公司权责手册2.0版（区域平台分册）9.5发各部门.xlsx"
OUT_PATH   = r"D:\西北投资\制度助理\data\authority.json"

# ── 第一步：用 openpyxl 展开合并格，构建完整表头 ──────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

# 把每个格子的值读到 expanded[row][col]
expanded = {}
for row in ws.iter_rows():
    for cell in row:
        expanded.setdefault(cell.row, {})[cell.column] = cell.value

# 展开合并格：把主格的值复制到每个合并子格
for rng in ws.merged_cells.ranges:
    val = expanded.get(rng.min_row, {}).get(rng.min_col)
    for r in range(rng.min_row, rng.max_row + 1):
        for c in range(rng.min_col, rng.max_col + 1):
            expanded.setdefault(r, {})[c] = val

def _str(v):
    s = str(v).strip() if v is not None else ''
    return '' if s == 'nan' else s

def get_col_label(openpyxl_col):
    """
    合并 row5（部门名）和 row6（条线子名）作为角色标签。
    openpyxl 列是 1-indexed，pandas 列是 0-indexed。
    """
    r5 = _str(expanded.get(5, {}).get(openpyxl_col))
    r6 = _str(expanded.get(6, {}).get(openpyxl_col))
    # 若 row5 空，回退到 row4、row3
    if not r5:
        for r in [4, 3, 2]:
            v = _str(expanded.get(r, {}).get(openpyxl_col))
            if v:
                r5 = v
                break
    if r5 and r6 and r6 != r5:
        return f"{r5}-{r6}"
    return r5 or r6

# pandas 0-indexed 列 → 角色标签
col_labels = {oc - 1: get_col_label(oc) for oc in range(5, 75)}

# ── 第二步：pandas 读数据行 ────────────────────────────────────────────────
df = pd.read_excel(EXCEL_PATH, sheet_name=0, header=None)

def get_role_name(pandas_col):
    return col_labels.get(pandas_col, '')

def get_org_level(col):
    if   4  <= col <= 17: return '项目公司'
    elif 18 <= col <= 29: return '西北投资本部'
    elif 30 <= col <= 38: return '西北投资领导'
    elif 39 <= col <= 42: return '西北投资决策'
    elif 43 <= col <= 67: return '中交投资总部'
    elif col == 68:        return '中国交建'
    return ''

STEP_ORDER = {'①':1,'②':2,'③':3,'④':4,'⑤':5,'⑥':6,'⑦':7,'⑧':8,'⑨':9}

items = []
for idx in range(6, len(df)):   # Excel 第7行起是数据（0-indexed idx=6）
    row = df.iloc[idx]
    key  = _str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
    name = _str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
    if not key or not name:
        continue

    def cv(i):
        return _str(row.iloc[i]) if pd.notna(row.iloc[i]) else ''

    initiator      = cv(2)
    final_approver = cv(3)
    system         = cv(69)
    frequency      = cv(70)
    is_authorized  = cv(71) == '是'
    remark         = cv(72)
    supplement_dept= cv(73)

    # 扫描流程列（pandas 列 4~68）
    numbered   = []
    countersign= []
    report     = []
    max_step   = 0

    for col in range(4, 69):
        cell = _str(row.iloc[col]) if pd.notna(row.iloc[col]) else ''
        if cell in STEP_ORDER and STEP_ORDER[cell] > max_step:
            max_step = STEP_ORDER[cell]

    for col in range(4, 69):
        cell = _str(row.iloc[col]) if pd.notna(row.iloc[col]) else ''
        if not cell:
            continue
        role = get_role_name(col)
        org  = get_org_level(col)

        if cell in STEP_ORDER:
            step_num = STEP_ORDER[cell]
            numbered.append({
                'step': cell,
                'step_num': step_num,
                'type': 'approve' if step_num == max_step else 'review',
                'is_final': step_num == max_step,
                'role': role,
                'org': org,
            })
        elif cell == '○':
            countersign.append({'step': '○', 'type': 'countersign', 'role': role, 'org': org})
        elif cell == '△':
            report.append({'step': '△', 'type': 'report', 'role': role, 'org': org})

    numbered.sort(key=lambda x: x['step_num'])

    # 排列顺序：① → 会签（并行） → ②③④... → △报备
    if numbered and countersign:
        flow = [numbered[0]] + countersign + numbered[1:] + report
    else:
        flow = countersign + numbered + report

    items.append({
        'key':            key,
        'name':           name,
        'initiator':      initiator,
        'final_approver': final_approver,
        'flow':           flow,
        'system':         system,
        'frequency':      frequency,
        'is_authorized':  is_authorized,
        'remark':         remark,
        'supplement_dept':supplement_dept,
    })

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, indent=2)

print(f'解析完成：{len(items)} 条权责事项 → {OUT_PATH}')

# 验证 R1-201
sample = next((i for i in items if i['key'] == 'R1-201'), None)
if sample:
    print(f'\n✓ R1-201：{sample["name"]}')
    for s in sample['flow']:
        final_mark = ' ★最终' if s.get('is_final') else ''
        print(f'  {s["step"]} [{s["type"]:12s}] {s["org"]:10s} / {s["role"]}{final_mark}')
